"""从 Excel / 飞书表格导入秋招岗位信息。"""
from __future__ import annotations

import datetime as dt
import hashlib
import io
import logging
import re
from pathlib import Path
from typing import Any, Optional

import openpyxl
from sqlalchemy.orm import Session

from ..config import BASE_DIR, get_settings
from ..models import Group, Job, get_sessionmaker

logger = logging.getLogger(__name__)

HYPERLINK_RE = re.compile(r'HYPERLINK\("([^"]+)"', re.IGNORECASE)
DATE_RE = re.compile(r"(\d{4})[.\-/年](\d{1,2})[.\-/月](\d{1,2})")

HEADER_MAP = {
    "公司": "company",
    "公司名称": "company",
    "公司类型": "company_type",
    "类型": "company_type",
    "批次": "batch",
    "base": "location",
    "地点": "location",
    "城市": "location",
    "岗位": "title",
    "岗位名称": "title",
    "职位": "title",
    "岗位jd": "description",
    "jd": "description",
    "岗位描述": "description",
    "投递链接": "url",
    "链接": "url",
    "开始日期": "open_date",
    "开放日期": "open_date",
    "截止日期": "close_date",
    "投递机制": "apply_rule",
    "内推码": "referrer_code",
    "记录时间": "recorded_at",
}


def uploads_dir() -> Path:
    path = BASE_DIR / "data" / "jobs"
    path.mkdir(parents=True, exist_ok=True)
    return path


def latest_upload_path(group_id: Optional[int] = None) -> Path:
    return uploads_dir() / (f"latest-{group_id}.xlsx" if group_id else "latest.xlsx")


def _resolve_excel_path(group_id: Optional[int] = None) -> Path:
    latest = latest_upload_path(group_id)
    if latest.exists():
        return latest
    settings = get_settings()
    configured = (settings.jobs.excel_path or "").strip()
    if configured:
        path = Path(configured)
        if not path.is_absolute():
            path = BASE_DIR / path
        return path
    return BASE_DIR.parent / "洒水车秋招_岗位信息表_总表.xlsx"


def _parse_url_value(val: Any, hyperlink_target: Optional[str] = None) -> Optional[str]:
    if hyperlink_target:
        return hyperlink_target
    if val is None:
        return None
    text = str(val).strip()
    if text.startswith("=HYPERLINK"):
        m = HYPERLINK_RE.search(text)
        return m.group(1) if m else None
    if text.startswith("http"):
        return text
    return None


def _parse_date(val: Any) -> Optional[dt.date]:
    if val is None:
        return None
    if isinstance(val, dt.datetime):
        return val.date()
    if isinstance(val, dt.date):
        return val
    text = str(val).strip()
    if not text or text in {"招满即止", "长期", "—", "-", "无"}:
        return None
    m = DATE_RE.search(text)
    if m:
        try:
            return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    try:
        return dt.date.fromisoformat(text[:10])
    except ValueError:
        return None


def _close_date_text(val: Any) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, (dt.datetime, dt.date)):
        return val.strftime("%Y-%m-%d")
    text = str(val).strip()
    return text or None


def _make_source_id(company: str, title: str, location: str, batch: str, url: str) -> str:
    raw = f"{company}|{title}|{location}|{batch}|{url}"
    return hashlib.md5(raw.encode()).hexdigest()


def _cell_text(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _normalize_header(name: Any) -> str:
    return _cell_text(name).lower().replace(" ", "")


def _row_dict_to_job(values: dict[str, Any], urls: dict[str, str], row_idx: int) -> Optional[dict]:
    company = _cell_text(values.get("company"))
    title = _cell_text(values.get("title"))
    if not company or not title:
        return None

    location = _cell_text(values.get("location")) or None
    batch = _cell_text(values.get("batch")) or None
    url = _parse_url_value(values.get("url"), urls.get("url"))
    close_raw = values.get("close_date")
    recorded_at = values.get("recorded_at")

    return {
        "source": "excel",
        "source_id": _make_source_id(company, title, location or "", batch or "", url or ""),
        "company": company,
        "title": title,
        "company_type": _cell_text(values.get("company_type")) or None,
        "batch": batch,
        "location": location,
        "description": _cell_text(values.get("description")) or None,
        "url": url,
        "open_date": _parse_date(values.get("open_date")),
        "close_date": _parse_date(close_raw),
        "close_date_text": _close_date_text(close_raw),
        "apply_rule": _cell_text(values.get("apply_rule")) or None,
        "referrer_code": _cell_text(values.get("referrer_code")) or None,
        "raw": {
            "row": row_idx,
            "recorded_at": recorded_at.isoformat() if isinstance(recorded_at, dt.datetime) else str(recorded_at or ""),
        },
    }


def _header_indexes(header_row: list[Any]) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for i, name in enumerate(header_row):
        mapped = HEADER_MAP.get(_normalize_header(name))
        if mapped and mapped not in indexes:
            indexes[mapped] = i
    return indexes


def parse_excel_workbook(content: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=False)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows())
    wb.close()
    if not rows:
        return []

    header_cells = rows[0]
    indexes = _header_indexes([c.value for c in header_cells])
    if "company" not in indexes or "title" not in indexes:
        # 兼容无表头、按洒水车默认列序
        indexes = {
            "company": 0,
            "company_type": 1,
            "batch": 2,
            "location": 3,
            "title": 4,
            "description": 5,
            "url": 6,
            "open_date": 7,
            "close_date": 8,
            "apply_rule": 9,
            "referrer_code": 10,
            "recorded_at": 11,
        }
        data_rows = rows[1:] if _cell_text(header_cells[0].value) in {"公司", "公司名称"} else rows
    else:
        data_rows = rows[1:]

    items: list[dict] = []
    for row_idx, row in enumerate(data_rows, start=2):
        if not any(c.value for c in row):
            continue
        values: dict[str, Any] = {}
        urls: dict[str, str] = {}
        for key, col in indexes.items():
            if col >= len(row):
                continue
            cell = row[col]
            values[key] = cell.value
            if getattr(cell, "hyperlink", None) and cell.hyperlink.target:
                urls[key] = cell.hyperlink.target
        parsed = _row_dict_to_job(values, urls, row_idx)
        if parsed:
            items.append(parsed)
    return items


def parse_value_rows(header: list[Any], data_rows: list[list[Any]]) -> list[dict]:
    indexes = _header_indexes(header)
    items: list[dict] = []
    for row_idx, row in enumerate(data_rows, start=2):
        if not any(v not in (None, "") for v in row):
            continue
        values = {}
        for key, col in indexes.items():
            if col < len(row):
                values[key] = row[col]
        parsed = _row_dict_to_job(values, {}, row_idx)
        if parsed:
            items.append(parsed)
    return items


def _find_existing(db: Session, it: dict) -> Optional[Job]:
    source = it.get("source", "excel")
    group_id = it.get("group_id")
    found = db.query(Job).filter_by(
        group_id=group_id, source=source, source_id=it["source_id"]
    ).first()
    if found:
        return found
    q = db.query(Job).filter(
        Job.source == source,
        Job.group_id == group_id,
        Job.company == it["company"],
        Job.title == it["title"],
    )
    if it.get("url"):
        by_url = q.filter(Job.url == it["url"]).first()
        if by_url:
            return by_url
    found = q.filter(Job.location == it.get("location"), Job.batch == it.get("batch")).first()
    return found if found else None


def _upsert_jobs(db: Session, items: list[dict], deactivate_missing: bool = False) -> dict:
    created = updated = 0
    seen: set[str] = set()
    for it in items:
        sid = it["source_id"]
        if sid in seen:
            continue
        seen.add(sid)
        existing = _find_existing(db, it)
        if existing:
            changed = existing.is_active is False
            for key in (
                "source", "source_id", "company", "title", "company_type", "batch", "location",
                "description", "url", "open_date", "close_date", "close_date_text",
                "apply_rule", "referrer_code", "raw",
            ):
                value = it.get(key)
                if getattr(existing, key) != value:
                    setattr(existing, key, value)
                    changed = True
            existing.is_active = True
            if changed:
                updated += 1
        else:
            db.add(Job(**it))
            created += 1
    deactivated = 0
    if deactivate_missing and items:
        source = items[0].get("source", "excel")
        group_id = items[0].get("group_id")
        stale = db.query(Job).filter(
            Job.source == source,
            Job.group_id == group_id,
            Job.is_active.is_not(False),
            ~Job.source_id.in_(seen),
        ).all()
        for job in stale:
            job.is_active = False
            deactivated += 1
    db.commit()
    return {"created": created, "updated": updated, "deactivated": deactivated, "total": len(items)}


def import_job_items(
    items: list[dict],
    db: Optional[Session] = None,
    source_label: str = "",
    deactivate_missing: bool = False,
    group_id: Optional[int] = None,
    created_by_id: Optional[int] = None,
) -> dict:
    own_session = db is None
    if own_session:
        db = get_sessionmaker()()
    try:
        source = source_label if source_label in {"excel", "manual"} else "excel"
        normalized = []
        for item in items:
            source_id = item["source_id"]
            if group_id and not source_id.startswith(f"g{group_id}-"):
                source_id = f"g{group_id}-{source_id}"
            normalized.append({
                **item,
                "source": source,
                "source_id": source_id,
                "group_id": group_id,
                "created_by_id": created_by_id,
            })
        result = _upsert_jobs(db, normalized, deactivate_missing=deactivate_missing)
        result["file"] = source_label
        logger.info("岗位导入完成: %s", result)
        return result
    finally:
        if own_session:
            db.close()


def import_jobs_from_bytes(
    content: bytes,
    save_as_latest: bool = True,
    db: Optional[Session] = None,
    group_id: Optional[int] = None,
    created_by_id: Optional[int] = None,
) -> dict:
    if save_as_latest:
        dest = latest_upload_path(group_id)
        dest.write_bytes(content)
        stamp = uploads_dir() / f"upload-{dt.datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
        stamp.write_bytes(content)
    items = parse_excel_workbook(content)
    if not items:
        raise ValueError("表格里没有识别到岗位，请确认表头包含「公司」「岗位」")
    return import_job_items(
        items, db=db, source_label="upload", group_id=group_id, created_by_id=created_by_id
    )


def import_jobs_from_excel(
    path: Optional[Path] = None,
    db: Optional[Session] = None,
    group_id: Optional[int] = None,
    created_by_id: Optional[int] = None,
) -> dict:
    excel_path = path or _resolve_excel_path(group_id)
    if not excel_path.exists():
        raise FileNotFoundError(f"岗位 Excel 不存在: {excel_path}")
    return import_jobs_from_bytes(
        excel_path.read_bytes(),
        save_as_latest=False,
        db=db,
        group_id=group_id,
        created_by_id=created_by_id,
    )


def import_jobs_from_config() -> dict:
    try:
        db = get_sessionmaker()()
        try:
            group = db.query(Group).filter(Group.is_system.is_(True)).order_by(Group.id).first()
            result = import_jobs_from_excel(db=db, group_id=group.id if group else None)
        finally:
            db.close()
        result["file"] = str(_resolve_excel_path())
        return result
    except FileNotFoundError as e:
        logger.warning(str(e))
        return {"created": 0, "updated": 0, "total": 0, "file": "", "skipped": True}
    except Exception:
        logger.exception("Excel 岗位导入失败")
        return {"created": 0, "updated": 0, "total": 0, "error": True}


def import_status(db: Optional[Session] = None, group_id: Optional[int] = None) -> dict:
    latest = latest_upload_path(group_id)
    configured = _resolve_excel_path(group_id)
    return {
        "excel_path": str(configured) if configured.exists() else "",
        "has_upload": latest.exists(),
        "uploaded_at": dt.datetime.fromtimestamp(latest.stat().st_mtime).isoformat() if latest.exists() else None,
    }
